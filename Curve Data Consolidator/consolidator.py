from openpyxl import load_workbook
from extracter import Info


def consolidate(file, extracted: list[Info]):
    workbook = load_workbook(file)
    sheet = workbook.create_sheet()
    
    titles = [
        "Curve No.",
        
        "Tangent Speed Max",
        "Tangent Speed Min",
        "Tangent Speed Mean",
        "Tangent Speed Mode",
        "Tangent HR Max",
        "Tangent HR Min",
        "Tangent HR Mean",
        "Tangent HR Mode",
        "Tangent GSR Max",
        "Tangent GSR Min",
        "Tangent GSR Mean",
        "Tangent GSR Mode",
        
        "CV Speed Max",
        "CV Speed Min",
        "CV Speed Mean",
        "CV Speed Mode",
        "CV HR Max",
        "CV HR Min",
        "CV HR Mean",
        "CV HR Mode",
        "CV GSR Max",
        "CV GSR Min",
        "CV GSR Mean",
        "CV GSR Mode",
        
        "Curve Speed Max",
        "Curve Speed Min",
        "Curve Speed Mean",
        "Curve Speed Mode",
        "Curve HR Max",
        "Curve HR Min",
        "Curve HR Mean",
        "Curve HR Mode",
        "Curve GSR Max",
        "Curve GSR Min",
        "Curve GSR Mean",
        "Curve GSR Mode"
    ]
    
    for i in range(1, len(titles)+1):
        sheet.cell(1, i, value=titles[i-1])
    
    for i in range(2, len(extracted)+2):
        info = extracted[i-2]
        
        sheet.cell(i, 1, value=info.curve_no)
        
        tangent_stables = info.tangent.get_stable_values()
        
        sheet.cell(i, 2, value=tangent_stables.speed_list.max())
        sheet.cell(i, 3, value=tangent_stables.speed_list.min())
        sheet.cell(i, 4, value=tangent_stables.speed_list.mean())
        sheet.cell(i, 5, value=tangent_stables.speed_list.mode())
        sheet.cell(i, 6, value=tangent_stables.hr_list.max())
        sheet.cell(i, 7, value=tangent_stables.hr_list.min())
        sheet.cell(i, 8, value=tangent_stables.hr_list.mean())
        sheet.cell(i, 9, value=tangent_stables.hr_list.mode())
        sheet.cell(i, 10, value=tangent_stables.gsr_list.max())
        sheet.cell(i, 11, value=tangent_stables.gsr_list.min())
        sheet.cell(i, 12, value=tangent_stables.gsr_list.mean())
        sheet.cell(i, 13, value=tangent_stables.gsr_list.mode())
        
        sheet.cell(i, 14, value=info.cv.speed_list.max())
        sheet.cell(i, 15, value=info.cv.speed_list.min())
        sheet.cell(i, 16, value=info.cv.speed_list.mean())
        sheet.cell(i, 17, value=info.cv.speed_list.mode())
        sheet.cell(i, 18, value=info.cv.hr_list.max())
        sheet.cell(i, 19, value=info.cv.hr_list.min())
        sheet.cell(i, 20, value=info.cv.hr_list.mean())
        sheet.cell(i, 21, value=info.cv.hr_list.mode())
        sheet.cell(i, 22, value=info.cv.gsr_list.max())
        sheet.cell(i, 23, value=info.cv.gsr_list.min())
        sheet.cell(i, 24, value=info.cv.gsr_list.mean())
        sheet.cell(i, 25, value=info.cv.gsr_list.mode())
        
        sheet.cell(i, 26, value=info.curve.speed_list.max())
        sheet.cell(i, 27, value=info.curve.speed_list.min())
        sheet.cell(i, 28, value=info.curve.speed_list.mean())
        sheet.cell(i, 29, value=info.curve.speed_list.mode())
        sheet.cell(i, 30, value=info.curve.hr_list.max())
        sheet.cell(i, 31, value=info.curve.hr_list.min())
        sheet.cell(i, 32, value=info.curve.hr_list.mean())
        sheet.cell(i, 33, value=info.curve.hr_list.mode())
        sheet.cell(i, 34, value=info.curve.gsr_list.max())
        sheet.cell(i, 35, value=info.curve.gsr_list.min())
        sheet.cell(i, 36, value=info.curve.gsr_list.mean())
        sheet.cell(i, 37, value=info.curve.gsr_list.mode())
    
    workbook.save(file)
    workbook.close()
