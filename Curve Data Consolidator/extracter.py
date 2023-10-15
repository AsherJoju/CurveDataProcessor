from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX
from road import Road, RoadState


class Data:
    
    def __init__(self, tangent_color, curve_color, speed, curve_no, hr, gsr):
        self.tangent_color = tangent_color
        self.curve_color = curve_color
        self.speed = speed
        self.curve_no = curve_no
        self.hr = hr
        self.gsr = gsr


class Info:
    
    def __init__(self):
        self.curve_no = ""
        self.tangent = Road()
        self.cv = Road()
        self.curve = Road()


class Colors:
    
    def __init__(self, tangent_color, cv_color, curve_color):
        self.tangent_color = tangent_color
        self.cv_color = cv_color
        self.curve_color = curve_color
    
    
    def set_color(file, sheet, cell):
        workbook = load_workbook(file)
        sheet = workbook.worksheets[sheet-1]
        
        color = Colors.get_color(sheet[cell].fill.start_color.index)
        
        workbook.close()
        
        return color


    def get_color(index: str):
        if str(index).isdecimal():
            return COLOR_INDEX[int(index)]
        return index


def extract(file, sheet, colors: Colors):
    workbook = load_workbook(file)
    sheet = workbook.worksheets[sheet-1]
    
    indices = [6]
    titles = list(sheet.iter_rows())[0]
    
    for i in range(len(titles)):
        if titles[i].value in [
            "LATITUDE",
            "SPEED(KM/H)",
            "HEART RATE",
            "GALVANIC SKIN RESPONSE"
        ]:
            indices.append(i)
    
    state = RoadState.NONE
    info_list: list[Info] = []
    info = Info()
    
    for row in sheet.iter_rows():
        if row[indices[1]].value == None:
            break
        
        data = Data(
            Colors.get_color(row[indices[1]].fill.start_color.index),
            Colors.get_color(row[indices[2]].fill.start_color.index),
            row[indices[2]].value,
            row[indices[0]].value,
            row[indices[3]].value,
            row[indices[4]].value
        )
        
        if data.curve_color == colors.curve_color:
            if state == RoadState.CURVE:
                if info.curve_no != "":
                    info_list.append(info)
                    info = Info()
                state = RoadState.NONE
            else:
                state = RoadState.CURVE
        if data.tangent_color == colors.tangent_color:
            state = RoadState.TANGENT
        elif data.tangent_color == colors.cv_color:
            state = RoadState.CV
        
        if state == RoadState.TANGENT:
            info.tangent.speed_list.values.append(data.speed)
            info.tangent.hr_list.values.append(data.hr)
            info.tangent.gsr_list.values.append(data.gsr)
        elif state == RoadState.CV:
            info.cv.speed_list.values.append(data.speed)
            info.cv.hr_list.values.append(data.hr)
            info.cv.gsr_list.values.append(data.gsr)
        elif state == RoadState.CURVE:
            if data.curve_no != None:
                info.curve_no = data.curve_no
            info.curve.speed_list.values.append(data.speed)
            info.curve.hr_list.values.append(data.hr)
            info.curve.gsr_list.values.append(data.gsr)
    
    workbook.close()
    
    return info_list
