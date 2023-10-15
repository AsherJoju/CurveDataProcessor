from openpyxl import load_workbook, Workbook
import os


def get_files(directory: str):
    dirs: list[str] = os.listdir(directory)
    
    files: dict[str, list[str]] = {}
    
    for dir in dirs:
        key: str = dir
        value = os.listdir(os.path.join(directory, dir))
        
        files[os.path.join(directory, key)] = []
    
        for file in value:
            files[os.path.join(directory, key)].append(
                os.path.join(directory, key, file)
            )
    
    return files


def copy_sheet(book: Workbook, sheet):
    sheet._parent = book # type: ignore
    book._add_sheet(sheet) # type: ignore
    return book


def collect(directory: str, sheet_name: str):
    files: dict[str, list[str]] = get_files(directory)
    out = Workbook()
    
    for key, value in files.items():
        name = os.path.join(key, str(os.path.basename(key)) + ".xlsx")
        
        book = Workbook()
        new = book.create_sheet(os.path.basename(key))
        
        for file in value:
            load = load_workbook(file)
            sheet = load[sheet_name]
            
            for row in sheet.iter_rows(values_only=True):
                new.append(row)
        
        book.save(name)
        out = copy_sheet(out, new)
    
    out.save(os.path.join(directory, str(os.path.basename(directory)) + ".xlsx"))
