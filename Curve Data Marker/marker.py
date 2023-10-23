from coordinates import Coordinates
from extracter import ExtractedData
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def mark(filename: str, sheetname: str, extracted_data: list[ExtractedData]) -> None:
    workbook = load_workbook(filename, data_only=True)
    worksheet = workbook[sheetname]
    
    column_indices: dict[str, int] = {}
    
    for title_row in worksheet.iter_rows(max_row=1):
        for title_cell in title_row:
            if title_cell.value in ["LATITUDE", "LONGITUDE", "CURVE NO"]:
                column_indices[str(title_cell.value)] = title_cell.column - 1 # Covert to 0-based
    
    for data in extracted_data:
        for value in data.values:
            min_row = None
            
            for row in worksheet.iter_rows(min_row=2):
                if row[column_indices["LATITUDE"]].value and row[column_indices["LONGITUDE"]].value:
                    pass
                else:
                    continue
                
                if min_row == None:
                    min_row = row
                    continue
                
                min_coordinates = Coordinates(
                    float(str(min_row[column_indices["LATITUDE"]].value)[:-1]),
                    float(str(min_row[column_indices["LONGITUDE"]].value)[:-1])
                )
                new_coordinates = Coordinates(
                    float(str(row[column_indices["LATITUDE"]].value)[:-1]),
                    float(str(row[column_indices["LONGITUDE"]].value)[:-1])
                )
                
                if new_coordinates.distance_to(value) < min_coordinates.distance_to(value):
                    min_row = row
        
            if min_row != None:
                if data.title == "CURVE START":
                    min_row[column_indices["CURVE NO"]].value = value.index
                for cell in min_row:
                    cell.fill = PatternFill(fill_type="solid", fgColor=data.color)
    
    workbook.save(filename)
