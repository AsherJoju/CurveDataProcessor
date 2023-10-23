from coordinates import Coordinates
from openpyxl import load_workbook


class ExtractedData:
    
    def __init__(self, title: str, color: str, column_index: int) -> None:
        self.title = title
        self.color = color
        self.column_index = column_index
        self.values: list[Coordinates] = []
    
    
    def __repr__(self) -> str:
        string = f"{self.column_index + 1}. {self.title} {self.color[2:]}\n"
        
        for value in self.values:
            string += f"\t{value}\n"
        
        return string


def extract(filename: str, sheetname: str) -> list[ExtractedData]:
    extracted_data: list[ExtractedData] = []
    
    workbook = load_workbook(filename, read_only=True, data_only=True)
    worksheet = workbook[sheetname]
    
    for title_row in worksheet.iter_rows(max_row=1):
        for title_cell in title_row:
            if title_cell.value in ["TANGENT", "CURVE VISIBILITY", "CURVE START", "CURVE END"]:
                extracted_data.append(
                    ExtractedData(
                        str(title_cell.value),
                        str(title_cell.fill.fgColor.rgb),
                        title_cell.column - 1 # Convert to 0-based
                    )
                )
    
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        for data in extracted_data:
            if row[data.column_index] and row[data.column_index + 1]:
                data.values.append(
                    Coordinates(
                        float(str(row[data.column_index])),
                        float(str(row[data.column_index + 1])),
                        str(row[0]) if data.title == "CURVE START" else ""
                    )
                )
    
    return extracted_data
